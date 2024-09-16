from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
import openpyxl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import pyautogui

options = Options()
options.add_argument("--new-tab")  # Open a new tab instead of a new window

# Initialize the WebDriver
driver = webdriver.Chrome(options=options)

# Open the web page
driver.get('http://bibaran.nvc.gov.np/sampati_fed/index.php')
time.sleep(4)

# Login
username_input = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.XPATH, "/html/body/div/center/div/div/div/div[2]/form/input[1]"))
)
password_input = driver.find_element(By.XPATH, "/html/body/div/center/div/div/div/div[2]/form/input[2]")
submit_button = driver.find_element(By.XPATH, "/html/body/div/center/div/div/div/div[2]/form/button")

username_input.send_keys('Enter username')  # Replace with your actual username
password_input.send_keys('enter password')  # Replace with your actual password
submit_button.click()

# Navigate to the target page
driver.get('http://bibaran.nvc.gov.np/sampati_fed/bibaran/bibaran.php')
workbook = load_workbook('E:\\Book_sampati.xlsx')  # Replace with your actual path
worksheet = workbook.active
time.sleep(2)

for row in worksheet.iter_rows(min_row=2, values_only=True):
    name, post, sex, office, mobile, dartasn, date, entry = row
    
    # Interact with form fields
    name_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/div[3]/div[2]/div/form/div/div[2]/div[2]/div/input"))
    )
    post_input = driver.find_element(By.XPATH, "/html/body/div[3]/div[2]/div/form/div/div[3]/div[1]/div/input")
    
    office_input = driver.find_element(By.XPATH, "/html/body/div[3]/div[2]/div/form/div/div[1]/div/div/input")
    mobile_input = driver.find_element(By.XPATH, "/html/body/div[3]/div[2]/div/form/div/div[4]/div[2]/div/input")
    dartasn_input = driver.find_element(By.XPATH, "/html/body/div[3]/div[2]/div/form/div/div[6]/div[1]/div/input")
    date_input = driver.find_element(By.XPATH, "/html/body/div[3]/div[2]/div/form/div/div[6]/div[2]/div/input")
    entry_input = driver.find_element(By.XPATH, "/html/body/div[3]/div[2]/div/form/div/div[7]/div[1]/div/input")
    ministry_dropdown = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "/html/body/div[3]/div[2]/div/form/div/div[8]/div/div/span/span[1]/span/span[1]"))
    )
    add_button = driver.find_element(By.XPATH, "/html/body/div[3]/div[2]/div/form/div/div[9]/button")

    # Scroll into view and set values in the form
    driver.execute_script("arguments[0].scrollIntoView(true);", name_input)
    name_input.send_keys(name)
    post_input.send_keys(post)
    office_input.send_keys(office)
    mobile_input.send_keys(mobile)
    dartasn_input.send_keys(dartasn)
    date_input.send_keys(date)
    entry_input.send_keys(entry)
    
    time_bujaune_input = driver.find_element(By.XPATH, "/html/body/div[3]/div[2]/div/form/div/div[2]/div[1]/div/span/span[1]/span")
    time_bujaune_input.click()
    time.sleep(2)  # Wait for dropdown animation    
    pyautogui.press('down')
    pyautogui.press('enter')
    
    pradesh_input = driver.find_element(By.XPATH, "/html/body/div[3]/div[2]/div/form/div/div[4]/div[1]/div/span/span[1]/span")
    pradesh_input.click()
    time.sleep(2)  # Wait for dropdown animation
    pyautogui.press('down')
    pyautogui.press('down')
    pyautogui.press('down')
    pyautogui.press('enter')
    
    
    
    nokari_input=driver.find_element(By.XPATH, "/html/body/div[3]/div[2]/div/form/div/div[5]/div[1]/div/span/span[1]/span/span[1]")
    nokari_input.click()
    time.sleep(2)  # Wait for dropdown animation
    pyautogui.press('down')
    pyautogui.press('enter')
    
    sex_input=driver.find_element(By.XPATH, "/html/body/div[3]/div[2]/div/form/div/div[3]/div[2]/div/span/span[1]/span")
    sex_input.click()
    time.sleep(2)  # Wait for dropdown animation
    pyautogui.press('enter')        #pyautogui.press('enter')
    job_type_input=driver.find_element(By.XPATH, "/html/body/div[3]/div[2]/div/form/div/div[5]/div[2]/div/span/span[1]/span/span[1]")
    job_type_input.click()
    time.sleep(2)
    pyautogui.press('down')
    pyautogui.press('enter')
    
    date_type_input=driver.find_element(By.XPATH, "/html/body/div[3]/div[2]/div/form/div/div[6]/div[2]/div/input")
    date_type_input.click()
    time.sleep(2)
    
    
    
    
    ministry_dropdown.click()
    time.sleep(30)
    

   
    
    time.sleep(1)
    # Add the entry
    add_button.click()
    time.sleep(4)
    #alert = WebDriverWait(driver, 10).until(EC.alert_is_present())
    #alert.dismiss()  # Closes the alert by pressing Cancel

    # Reload the form page
    driver.get('http://bibaran.nvc.gov.np/sampati_fed/bibaran/bibaran.php')
    time.sleep(4)

# Clean up
driver.quit()
